import datetime
import matplotlib.pyplot as plt
import numpy as np
import os
from openpyxl import load_workbook, Workbook

class GroceryManagementSystem:
    def __init__(self, file_name):
        self.file_name = file_name
        self.data = {
            "ichimliklar": {},
            "oziq-ovqat": {},
            'mevalar': {}
        }
        self.sell_data = {
            "ichimliklar": {},
            "mevalar": {},
            "oziq-ovqat": {},
        }
        self.add_data = {
            "ichimliklar": {},
            "mevalar": {},
            "oziq-ovqat": {},
        }

    def database_excel_drinks(self, category):
        global name_list_drinks, quenty_list_drinks, price_list_drinks, date_time_list_drinks, expired_time_list_drinks, country_list_drinks
        name_list_drinks = []
        quenty_list_drinks = []
        price_list_drinks = []
        date_time_list_drinks = []
        expired_time_list_drinks = []
        country_list_drinks = []
        wb = load_workbook('ali.xlsx')
        data_type = wb[category]
        for i in range(2, data_type.max_row + 1):
            name_list_drinks.append(data_type.cell(row=i, column=1).value)
            quenty_list_drinks.append(data_type.cell(row=i, column=2).value)
            price_list_drinks.append(data_type.cell(row=i, column=3).value)
            date_time_list_drinks.append(data_type.cell(row=i, column=4).value)
            expired_time_list_drinks.append(data_type.cell(row=i, column=5).value)
            country_list_drinks.append(data_type.cell(row=i, column=6).value)

        for j in name_list_drinks:
            self.data['ichimliklar'][j] = {
                "miqdor": quenty_list_drinks[name_list_drinks.index(j)],
                "price": price_list_drinks[name_list_drinks.index(j)],
                "date": date_time_list_drinks[name_list_drinks.index(j)],
                "sroki": expired_time_list_drinks[name_list_drinks.index(j)],
                "country": country_list_drinks[name_list_drinks.index(j)],
            }

        return self.data

    def database_excel_fruits(self, category):
        global name_list_fruits, quenty_list_fruits, price_list_fruits, date_time_list_fruits, expired_time_list_fruits, country_list_fruits
        name_list_fruits = []
        quenty_list_fruits = []
        price_list_fruits = []
        date_time_list_fruits = []
        expired_time_list_fruits = []
        country_list_fruits = []
        wb = load_workbook('ali.xlsx')
        data_type = wb[category]
        for i in range(2, data_type.max_row + 1):
            name_list_fruits.append(data_type.cell(row=i, column=1).value)
            quenty_list_fruits.append(data_type.cell(row=i, column=2).value)
            price_list_fruits.append(data_type.cell(row=i, column=3).value)
            date_time_list_fruits.append(data_type.cell(row=i, column=4).value)
            expired_time_list_fruits.append(data_type.cell(row=i, column=5).value)
            country_list_fruits.append(data_type.cell(row=i, column=6).value)

        for j in name_list_fruits:
            self.data['mevalar'][j] = {
                "miqdor": quenty_list_fruits[name_list_fruits.index(j)],
                "price": price_list_fruits[name_list_fruits.index(j)],
                "date": date_time_list_fruits[name_list_fruits.index(j)],
                "sroki": expired_time_list_fruits[name_list_fruits.index(j)],
                "country": country_list_fruits[name_list_fruits.index(j)],
            }

    def database_excel_food(self, category):
        global name_list_food, quenty_list_food, price_list_food, date_time_list_food, expired_time_list_food, country_list_food
        name_list_food = []
        quenty_list_food = []
        price_list_food = []
        date_time_list_food = []
        expired_time_list_food = []
        country_list_food = []
        wb = load_workbook('ali.xlsx')
        data_type = wb[category]
        for i in range(2, data_type.max_row + 1):
            name_list_food.append(data_type.cell(row=i, column=1).value)
            quenty_list_food.append(data_type.cell(row=i, column=2).value)
            price_list_food.append(data_type.cell(row=i, column=3).value)
            date_time_list_food.append(data_type.cell(row=i, column=4).value)
            expired_time_list_food.append(data_type.cell(row=i, column=5).value)
            country_list_food.append(data_type.cell(row=i, column=6).value)
        for j in name_list_food:
            self.data['oziq-ovqat'][j] = {
                "miqdor": quenty_list_food[name_list_food.index(j)],
                "price": price_list_food[name_list_food.index(j)],
                "date": date_time_list_food[name_list_food.index(j)],
                "sroki": expired_time_list_food[name_list_food.index(j)],
                "country": country_list_food[name_list_food.index(j)],
            }

    def add_product(self, *category):
        while True:
            print(" Meva qo'shish >> 1 \n"
                  " Ichimlik  qo'shish >> 2\n"
                  " oziq-ovqat qo'shish >> 3\n"
                  " Bo'limdan chiqish >> 4\n")
            sorov = (input())
            if sorov == '4':
                break
            elif sorov == '1':
                category = 'mevalar'
            elif sorov == '2':
                category = 'ichimliklar'
            elif sorov == '3':
                category = 'oziq-ovqat'

            product_name = input("Mahsulot nomini kiriting:\n")
            if product_name in self.data[category].keys():
                request_fruit = input(f"{product_name} dan bazada {self.data[category][product_name]['miqdor']} bor: \n"
                                      "Yana qo'shish uchun 1 ni bosing\n "
                                      "Qo'shmoqchi bo'lmasangiz 0 ni bosing\n")
                if request_fruit == '1':
                    while True:
                        weight_fruit = input(f"{category} og'irligini kiriting:\n")
                        try:
                            weight_fruit = float(weight_fruit)
                            if weight_fruit > 0:
                                self.data[category][product_name]['miqdor'] += weight_fruit
                                if product_name not in self.add_data[category]:
                                    self.add_data[category][product_name] = []
                                self.add_data[category][product_name].append({
                                    'miqdori': weight_fruit,
                                    'price': self.data[category][product_name]['price'],
                                    'date': datetime.datetime.now().strftime("%x"),
                                })
                                break
                            else:
                                print("Iltimos miqdorga musbat son kiriting:\n")
                        except:
                            print("Iltimos miqdorni to'g'ri kiriting")
            else:
                while True:
                    weight = input(f"{product_name} ning miqdorini kiriting:\n")
                    try:
                        weight = float(weight)
                        if weight > 0:
                            break
                        else:
                            print("Iltimos miqdorga musbat son kiriting:\n")
                    except:
                        print("Iltimos miqdorga to'g'ri qiymat kiriting:\n")
                while True:
                    price = input(f"{product_name} narxini kiriting:\n")
                    try:
                        price = float(price)
                        if price > 0:
                            break
                        else:
                            print("Iltimos narxga musbat son kiriting:\n")
                    except:
                        print("Iltimos narxga to'g'ri qiymat kiriting")
                add_date_time = datetime.datetime.now().strftime("%x")
                expiration_date = input(
                    f"{product_name} ning amal qilish muddatini kiriting(dd/mm/yyyy) ko'rinishida\n")
                country = input("Ishlab chiqarish mamlakatini kiriting\n")
                self.data[category][product_name] = {
                    "miqdor": weight,
                    "price": price,
                    "date": datetime.datetime.now().strftime("%x"),
                    "sroki": expiration_date,
                    "country": country,
                }
                if product_name not in self.add_data[category]:
                    self.add_data[category][product_name] = []
                self.add_data[category][product_name].append({
                    'miqdori': weight,
                    'price': price,
                    'date': datetime.datetime.now().strftime("%x")
                })
        self.main()

    def sell_product(self, *category):
        chek_name = []
        chek_weight = []
        chek_price = []
        while True:
            sorov_sell = input(" Meva tanlash >> 1 \n"
                               " Ichimlik tanlash >> 2 \n"
                               " Oziq-ovqat tanlash >> 3 \n"
                               " Tanlangan mahsulotlarni sotib olish >> 4 \n"
                               " Savatni tozalash >> 5 \n"
                               "Sotuv bo'limidan chiqish uchun tugmalardan birini bosing🤖!\n\n")
            if sorov_sell in ['1', '2', '3']:
                if sorov_sell == '1':
                    category = 'mevalar'
                elif sorov_sell == '2':
                    category = 'ichimliklar'
                elif sorov_sell == '3':
                    category = 'oziq-ovqat'
                product_sell_name = input("Mahsulot nomini kiriting:\n")
                if product_sell_name in self.data[category].keys():
                    while True:
                        fruit_sell_weight = input(f"{product_sell_name} ning miqdorini kiriting:\n")
                        try:
                            fruit_sell_weight = float(fruit_sell_weight)
                            if fruit_sell_weight > 0 and self.data[category][product_sell_name][
                                'miqdor'] >= fruit_sell_weight:
                                break
                            else:
                                print("Iltimos miqdorga to'g'ri qiymat kiriting kiriting: \n")
                        except ValueError:
                            print("Iltimos miqdorga yaroqli qiymat kiriting")
                    if product_sell_name not in self.sell_data[category]:
                        self.sell_data[category][product_sell_name] = []
                    self.sell_data[category][product_sell_name].append({
                        'miqdor': fruit_sell_weight,
                        'price': self.data[category][product_sell_name]['price'],
                        'date': datetime.datetime.now().strftime("%x")
                    })
                    chek_name.append(product_sell_name)
                    chek_weight.append(fruit_sell_weight)
                    chek_price.append(self.data[category][product_sell_name]['price'])
                    print(f"{product_sell_name} savatga qo'shildi!")
                    self.data[category][product_sell_name]['miqdor'] -= fruit_sell_weight
                    if self.data[category][product_sell_name]['miqdor'] == 0:
                        del self.data[category][product_sell_name]

                else:
                    print('Mavjud bolmagan mahsulotni tanladinggiz')

            elif sorov_sell == '4':
                if len(chek_name) == 0:
                    print("Mahsulot tanlashinggizni sorayman")
                else:
                    overal = 0
                    print("--->>>---🎉Haridingiz uchun Raxmat🎉-----<<<-----")
                    for i, k in enumerate(chek_name):
                        print(
                            f"{k}-----------------------{chek_weight[i]}x{chek_price[i]}={chek_weight[i] * chek_price[i]}")
                        price_sell = chek_weight[i] * chek_price[i] * 0.05
                        overal += chek_weight[i] * chek_price[i] + price_sell
                    print(f"JAMI ------------------------ {overal}\n"
                          f"TIME -------------------------{datetime.datetime.now()}")
                    chek_name.clear()
                    chek_price.clear()
                    chek_weight.clear()
            elif sorov_sell == '5':
                chek_name.clear()
                chek_price.clear()
                chek_weight.clear()
                print("Xaridlar tarixi tozalandi!\n\n")

            else:
                print("Xarid bo'limidan chiqdingiz\n\n")
                self.main()

    def display_report(self):
        def display_report_data(data):
            while True:
                date_1 = input("Sanani kiriting (MM/DD/YY ko'rinishda)\n"
                               "[0] Oldingi menyuga qaytish\n")
                if date_1 == '0':
                    break
                product_quantities = {}
                for category, products in data.items():
                    for product, values in products.items():
                        for value in values:
                            if value['date'] == date_1:
                                quantity = value['miqdori']
                                if product in product_quantities:
                                    product_quantities[product] += quantity
                                else:
                                    product_quantities[product] = quantity

                for product, quantity in product_quantities.items():
                    print(f"Maxsulot: {product}------------------ Miqdori: {quantity}\n\n")

        while True:
            request_report = input(" Bazadagi mahsulotlarni ko'rish >> 1\n"
                                   " Qo'shilgan mahsulotlarni ko'rish >> 2\n"
                                   " Sotilgan mahsulotlarni ko'rish >> 3\n"
                                   " Oldingni menyuga qaytish >> 4\n")
            if request_report == '1':
                array_1 = []
                array_1.append(len(self.data['mevalar']))
                array_1.append(len(self.data['ichimliklar']))
                array_1.append(len(self.data['oziq-ovqat']))
                y = np.array(array_1)
                mylabels = [f"mevalar {array_1[0]}", f"ichimliklar {array_1[1]}", f"oziq-ovqat {array_1[2]}"]
                plt.pie(y, labels=mylabels, startangle=90)
                plt.show()
                file_path = "/home/gofa/ICT_tasks/openpyxl/final_project/ali.xlsx"
                if os.path.exists(file_path):
                    os.system("libreoffice " + file_path)
                else:
                    print("Fayl topilmadi")
                while True:
                    request_report_product = input(" Mahsulot kiritish >> 1\n"
                                                   " Mahsulot sotish >> 2 \n"
                                                   " Hisobotlarni ko'rish >> 3\n"
                                                   " Dasturdan chiqish >> 4\n\n")

                    def display_report_product(category):
                        print(f"----->>>---Bazada mavjud {category}---<<<-----")
                        for i in self.data[category].keys():
                            print(f"{i}--------------------------------{self.data[category][i]['miqdor']} kg")

                    if request_report_product == '1':
                        display_report_product(category='mevalar')
                    elif request_report_product == '2':
                        display_report_product(category='ichimliklar')
                    elif request_report_product == '3':
                        display_report_product(category='oziq-ovqat')
                    elif request_report_product == '4':
                        break
            elif request_report == '2':
                display_report_data(data=self.add_data)
            elif request_report == '3':
                display_report_data(data=self.sell_data)
            elif request_report == '4':
                self.main()
    def main(self):
        print(" Mahsulot kiritish >> 1\n"
              " Mahsulot sotish >> 2 \n"
              "Hisobotlarni ko'rish >> 3\n"
              " Dasturdan chiqish >> 4\n")
        sorov = input(">>>")
        if sorov == '1':
            self.add_product()
        elif sorov == '2':
            self.sell_product()
        elif sorov == '3':
            self.display_report()
        elif sorov == '4':
            print("Dastur yakunlandi👌")
            wb = Workbook()

            # Iterate over each category
            for category, items in self.data.items():
                # Create a new worksheet for each category
                ws = wb.create_sheet(title=category)
                # Write headers
                headers = ["Nomi", "Miqdori", "Narxi", "Kelgan_sana", "Saqlash_muddati", "from_country"]
                ws.append(headers)
                # Write data for each item
                for name, details in items.items():
                    ws.append(
                        [name, details["miqdor"], details["price"], details["date"], details["sroki"],
                         details["country"]])

            # Remove the default sheet created by openpyxl
            wb.remove(wb["Sheet"])

            # Save the workbook
            wb.save("ali.xlsx")
            quit()
        self.main()




if __name__ == "__main__":
    file_name = input("File nomini kiriting:")+".xlsx"
    gms = GroceryManagementSystem(file_name)
    gms.database_excel_food(category='oziq-ovqat')
    gms.database_excel_fruits(category='mevalar')
    gms.database_excel_drinks(category='ichimliklar')

    gms.main()
